from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
import os
import datetime

## data merging
def transferWSData(ws):
    data = []
    num_row = len(ws['A'])
    num_col = len(ws[1])

    for row in ws.iter_rows(min_row=2,max_col=num_col,max_row=num_row):
        row_list = []
        for i in range(0,num_col):
            row_list.append(row[i].value)
        data.append(row_list)
    return data

def appendData(data, new_ws):
    for row in data:
        new_ws.append(row)

def merge_data(directory, filename):
    path = os.path.join(directory, filename)
    wb = load_workbook(path)

    new_wb = Workbook()
    new_invoice_ws = new_wb.create_sheet("invoice")
    new_invoice_ws.append(["Invoice_ID","Customer_ID", "Date_issued"])
    new_customer_ws = new_wb.create_sheet("customer")
    new_customer_ws.append(["ID","FirstName", "LastName", "Email", "Phone#"])
    new_invoice_list_ws = new_wb.create_sheet("invoice_list")
    new_invoice_list_ws.append(["Item_ID","Invoice_ID","Quantity"])
    new_product_ws = new_wb.create_sheet("product")
    new_product_ws.append(["Item_ID", "Item_name", "Item_price"])
    new_wb.remove(new_wb['Sheet'])
    print(new_wb.sheetnames)

    invoice_ws = wb["invoice"]
    customer_ws = wb["customer"]
    invoice_list_ws = wb["invoice_list"]
    product_ws = wb["product"]
    
    invoice_data = transferWSData(invoice_ws)
    customer_data = transferWSData(customer_ws)
    invoice_list_data = transferWSData(invoice_list_ws)
    product_data = transferWSData(product_ws)

    #append the appropriate table
    appendData(invoice_data, new_invoice_ws)
    appendData(customer_data, new_customer_ws)
    appendData(invoice_list_data, new_invoice_list_ws)
    appendData(product_data, new_product_ws)

    new_wb.save("merged.xlsx")

## data validation
def validateWB(directory, filename):
    ws_list = ["customer", "invoice", "invoice_list", "product"]
    missing_ws_list = []
    path = os.path.join(directory, filename)
    wb = load_workbook(path)

    try:
        for name in ws_list:
            if name in wb.sheetnames:
                continue
            else:
                missing_ws_list.append(name)
        if len(missing_ws_list) != 0:
            raise EnvironmentError
    except EnvironmentError:
        print("Data is invalid due to the following reason:")
        print(f'Missing WorkSheet:')
        string = ''
        for ws in missing_ws_list:
            string += f'{ws}'
        print(f"Can not find the following worksheet/s: '{string}'")
    except:
        print("Directory path is invalid")
    
    invoice_ws = wb["invoice"]
    customer_ws = wb["customer"]
    invoice_list_ws = wb["invoice_list"]
    product_ws = wb["product"]

    validateWS(customer_ws, "customer")
    validateWS(invoice_ws, "invoice")
    validateWS(invoice_list_ws, "invoice_list")
    validateWS(product_ws, "product")

    wb.save(path)

def validateWS(ws, wsName):
    num_row = len(ws['A'])
    intDV = DataValidation(type="whole")
    nameDV = DataValidation(type="textLength", operator="lessThanOrEqual", formula1=100)
    emailDV = DataValidation(type="custom", formula1='ISNUMBER(MATCH("*@*.?*",D2,0))')
    phoneDV = DataValidation(type="whole", operator="between", formula1=1000000000, formula2=9999999999)
    dateDV = DataValidation(type="date", operator="lessThan", formula1='DATEVALUE("2021/1/1")')
    if wsName == "customer":
        intDV.add(f'A2:A{num_row}')
        nameDV.add(f'B2:C{num_row}')
        emailDV.add(f'D2:D{num_row}')
        phoneDV.add(f'E2:E{num_row}')
        ws.add_data_validation(intDV)
        ws.add_data_validation(nameDV)
        ws.add_data_validation(emailDV)
        ws.add_data_validation(phoneDV)
    elif wsName == "invoice":
        intDV.add(f'A2:B{num_row}')
        dateDV.add(f'C2:C{num_row}')
        ws.add_data_validation(intDV)
        ws.add_data_validation(dateDV)
    elif wsName == "invoice_list":
        intDV.add(f'A2:C{num_row}')
        ws.add_data_validation(intDV)
    elif wsName == "product":
        nameDV.add(f'B2:B{num_row}')
        ws.add_data_validation(nameDV)
        intDV.add(f'A2:A{num_row}')
        intDV.add(f'C2:C{num_row}')
        ws.add_data_validation(intDV)

## insert data into dict
wb = load_workbook('./data_tables/shopdata.xlsx')

invoice_ws = wb["invoice"]
customer_ws = wb["customer"]
invoice_list_ws = wb["invoice_list"]
product_ws = wb["product"]

customer_dict = {}
product_dict = {}
invoice_dict = {}
invoice_list_dict = {}

def storeWbIntoDict(ws, targetDict):
    num_row = len(ws['A'])
    num_col = len(ws[1])
    header_list = ws[1]

    for row in ws.iter_rows(min_row=2,max_col=num_col,max_row=num_row):
        row_obj = {}
        for i in range(1,num_col):
            row_obj[header_list[i].value] = row[i].value

        targetDict[row[0].value] = row_obj

def storeInvoiceList(ws, targetDict):
    num_row = len(ws['A'])
    num_col = len(ws[1])
    header_list = ws[1]

    for row in ws.iter_rows(min_row=2,max_col=num_col,max_row=num_row):
        invoice_key = row[1].value
        if invoice_key in targetDict:
            targetDict[invoice_key].append({header_list[0].value: row[0].value, header_list[2].value: row[2].value})
        else:
            targetDict[invoice_key] = [{header_list[0].value: row[0].value, header_list[2].value: row[2].value}]


storeWbIntoDict(customer_ws,customer_dict)
storeWbIntoDict(product_ws,product_dict)
storeWbIntoDict(invoice_ws,invoice_dict)
storeInvoiceList(invoice_list_ws, invoice_list_dict)

## create invoice from dictionaries

def createInvoice(invoiceID):
    if invoiceID not in invoice_dict:
        return "invoice does not exist"

    customerID = invoice_dict[invoiceID]["Customer_ID"]
    date = invoice_dict[invoiceID]['Date_issued']
    firstName = customer_dict[customerID]['FirstName']
    lastName = customer_dict[customerID]['LastName']
    email = customer_dict[customerID]['Email']
    invoice_item = invoice_list_dict[invoiceID]

    invoice_item_chart = []
    total = 0
    for item in invoice_item: 
        itemName = product_dict[item['Item_ID']]['Item_name']
        itemUnitPrice = product_dict[item['Item_ID']]['Item_price']
        quantity = item['Quantity']
        total = total + itemUnitPrice * quantity
        invoice_item_chart.append({"Item":itemName, "Quantity": quantity, "Price": itemUnitPrice, "Amount": quantity * itemUnitPrice})
    
    with open('invoice.txt', mode='w') as invoice:
        invoice.write(f'Invoice For:\n {firstName} {lastName} \n {email} \n \n')
        invoice.write(f'Invoice ID: {invoiceID} \n Sent: {date.date()} \n Due: {(date + datetime.timedelta(days=14)).date()} \n \n' )
        invoice.write(f"ITEM            Quantity    Unit Price      Amount \n \n")
        for item in invoice_item_chart:
            invoice.write(f'{item["Item"]}          {item["Quantity"]}        ${item["Price"]}.00            ${item["Amount"]}.00 \n \n')
        invoice.write(f"SUBTOTAL:${total}.00\n TAX:${round(total * 0.1, 2)}")

if __name__ == '__main__':
    print('.....validating data.....')
    invoice_id = input("ID of the invoice you want to print:  ")
    validateWB("./data_tables", "shopdata.xlsx")
    print(".....creating invoice.....")
    createInvoice(int(invoice_id))
    print("***invoice successfully created***")
    # merge_data("./data_tables", "shopdata.xlsx")


