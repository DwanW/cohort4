import pytest
import os
from excel import *
from openpyxl import load_workbook

@pytest.fixture
def mock_wb():
    from openpyxl import Workbook

    mock_customer_data = [1,"Dwan", "Wang", "Dwan.W@gmail.com", 1234567890]
    mock_invoice_data = [3, 1, datetime.datetime(2020, 6, 6, 0, 0)]
    mock_invoice_list_data = [1,3,10]
    mock_product_data = [1, 'awesome code', 99]

    new_wb = Workbook()
    new_invoice_ws = new_wb.create_sheet("invoice")
    new_invoice_ws.append(["Invoice_ID","Customer_ID", "Date_issued"])
    new_customer_ws = new_wb.create_sheet("customer")
    new_customer_ws.append(["ID","FirstName", "LastName", "Email", "Phone#"])
    new_invoice_list_ws = new_wb.create_sheet("invoice_list")
    new_invoice_list_ws.append(["Item_ID","Invoice_ID","Quantity"])
    new_product_ws = new_wb.create_sheet("product")
    new_product_ws.append(["Item_ID", "Item_name", "Item_price"])
    new_wb.remove(new_wb['Sheet'])

    new_invoice_ws.append(mock_invoice_data)
    new_customer_ws.append(mock_customer_data)
    new_invoice_list_ws.append(mock_invoice_list_data)
    new_product_ws.append(mock_product_data)

    return new_wb
    
## data merging test
def test_transferWSData(mock_wb):
    mock_customer_ws = mock_wb["customer"]
    mock_invoice_ws = mock_wb["invoice"]
    assert transferWSData(mock_customer_ws) == [[1,"Dwan", "Wang", "Dwan.W@gmail.com", 1234567890]]
    assert transferWSData(mock_invoice_ws) == [[3, 1, datetime.datetime(2020, 6, 6, 0, 0)]]

def test_appendData(mock_wb):
    mock_data = [[2,"testfn", "testln", "test@gmail.com", 1334557990]]
    mock_customer_ws = mock_wb["customer"]

    appendData(mock_data, mock_customer_ws)
    assert len(mock_customer_ws['A']) == 3
    appendData(mock_data, mock_customer_ws)
    assert len(mock_customer_ws['A']) == 4

def test_mergeData(mock_wb):
    mock_wb.save("test.xlsx")
    directory = ""
    filename = "test.xlsx"

    merge_data(directory,filename)
    assert os.path.exists('merged.xlsx') == True
    generated_wb = load_workbook("merged.xlsx")
    assert generated_wb.sheetnames == ['invoice','customer', 'invoice_list', 'product']

## data validation test
def test_validateWS():
    pass

def test_validateWB():
    pass

## invoice creation test
def test_storeWSIntoDict(mock_wb):
    customer_ws = mock_wb["customer"]
    product_ws = mock_wb["product"]
    invoice_list_ws = mock_wb["invoice_list"]
    
    customer_dict = {}
    product_dict = {}
    invoice_list_dict = {}

    storeWSIntoDict(customer_ws, customer_dict)
    storeWSIntoDict(product_ws, product_dict)
    assert customer_dict == {1: {'Email': 'Dwan.W@gmail.com','FirstName': 'Dwan','LastName': 'Wang','Phone#': 1234567890}}
    assert product_dict == {1: {'Item_name': 'awesome code', 'Item_price': 99}}
    
    storeInvoiceList(invoice_list_ws, invoice_list_dict)
    assert invoice_list_dict == {3: [{'Item_ID': 1, 'Quantity': 10}]}

def test_createInvoice():
    assert createInvoice(2, "test.xlsx") == 'invoice does not exist'
    assert createInvoice(3, "test.xlsx") == 'invoice created'

    count = 0
    with open('invoice.txt') as invoice:
        for line in invoice:
            count += line.count('Invoice ID: 3')
            count += line.count('Dwan')
        assert count == 3



